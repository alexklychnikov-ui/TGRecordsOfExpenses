import os
import sqlite3
from typing import Optional, List, Tuple
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers


def _fetch_rows(db_path: str, username: str, start_date: Optional[str] = None, end_date: Optional[str] = None) -> Tuple[List[str], List[Tuple]]:
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        if start_date and end_date:
            cur.execute(
                "SELECT * FROM purchases WHERE date >= ? AND date <= ? AND username = ? ORDER BY date DESC, id ASC",
                (start_date, end_date, username),
            )
        else:
            cur.execute(
                "SELECT * FROM purchases WHERE username = ? ORDER BY date DESC, id ASC",
                (username,),
            )
        rows = cur.fetchall()
        columns = [d[0] for d in cur.description]
        return columns, rows
    finally:
        conn.close()


def _is_price_column(col_name: str) -> bool:
    return col_name in {"price", "discount"}


def _is_date_column(col_name: str) -> bool:
    return col_name in {"date", "created_at"}


def _coerce_cell_value(col: str, val):
    if val is None:
        return None
    if _is_price_column(col):
        try:
            return float(val)
        except Exception:
            return val
    if _is_date_column(col):
        # Return parsed datetime when possible; leave string otherwise
        try:
            if val is None:
                return None
            s = str(val)
            # dd.mm.yyyy
            if "." in s and len(s) >= 10:
                return datetime.strptime(s[:10], "%d.%m.%Y")
            # ISO 8601
            s2 = s.replace("Z", "+00:00")
            return datetime.fromisoformat(s2[:26])
        except Exception:
            return str(val)
    # Try int/float otherwise keep as-is
    if isinstance(val, (int, float)):
        return val
    try:
        if "." in str(val) or "," in str(val):
            return float(str(val).replace(",", "."))
        return int(val)
    except Exception:
        return val


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                v = cell.value
                s = v if isinstance(v, str) else ("" if v is None else str(v))
                if len(s) > max_len:
                    max_len = len(s)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="DDDDDD")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_body_style(cell, col_name: str) -> None:
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if _is_price_column(col_name):
        cell.number_format = numbers.FORMAT_NUMBER_00
    elif _is_date_column(col_name):
        # Ensure date formatting dd.mm.yyyy
        try:
            if isinstance(cell.value, datetime):
                cell.number_format = "dd.mm.yyyy"
            else:
                # Try parse strings once more
                v = str(cell.value)
                dt = None
                if "." in v and len(v) >= 10:
                    dt = datetime.strptime(v[:10], "%d.%m.%Y")
                elif "-" in v and len(v) >= 10:
                    # handle YYYY-MM-DD
                    dt = datetime.strptime(v[:10], "%Y-%m-%d")
                if dt:
                    cell.value = dt
                    cell.number_format = "dd.mm.yyyy"
                else:
                    cell.number_format = "@"
        except Exception:
            cell.number_format = "@"


_RU_HEADERS = {
    "id": "id (идентификатор записи)",
    "chequeid": "номер чека",
    "file_path": "путь к файлу фото",
    "date": "дата чека",
    "created_at": "дата создания записи",
    "product_name": "наименование продукта",
    "quantity": "количество",
    "price": "цена",
    "discount": "скидка",
    "category1": "категория 1",
    "category2": "категория 2",
    "category3": "категория 3",
    "organization": "организация",
    "username": "пользователь",
    "description": "комментарий",
}


def export_to_excel(db_path: str, output_path: str, username: str, start_date: Optional[str] = None, end_date: Optional[str] = None) -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    columns, rows = _fetch_rows(db_path, username, start_date, end_date)

    # Исключаем из вывода: file_path, created_at, discount, username
    excluded_columns = {"file_path", "created_at", "discount", "username"}
    filtered_columns = [col for col in columns if col not in excluded_columns]
    column_indices = [i for i, col in enumerate(columns) if col not in excluded_columns]
    
    # Перемещаем quantity сразу после product_name
    if "product_name" in filtered_columns and "quantity" in filtered_columns:
        product_idx = filtered_columns.index("product_name")
        quantity_idx = filtered_columns.index("quantity")
        if quantity_idx != product_idx + 1:
            # Удаляем quantity из текущей позиции
            filtered_columns.pop(quantity_idx)
            orig_quantity_idx = column_indices.pop(quantity_idx)
            # Вставляем quantity после product_name
            filtered_columns.insert(product_idx + 1, "quantity")
            column_indices.insert(product_idx + 1, orig_quantity_idx)

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchases"

    # Header row with RU descriptions
    for col_idx, col_name in enumerate(filtered_columns, start=1):
        header_text = _RU_HEADERS.get(col_name, col_name)
        c = ws.cell(row=1, column=col_idx, value=header_text)
        _apply_header_style(c)

    # Freeze header
    ws.freeze_panes = "A2"

    # Body rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, orig_idx in enumerate(column_indices, start=1):
            col_name = filtered_columns[c_idx - 1]
            val = row[orig_idx]
            coerced = _coerce_cell_value(col_name, val)
            cell = ws.cell(row=r_idx, column=c_idx, value=coerced)
            _apply_body_style(cell, col_name)

    # Autofilter over full data range
    last_row = ws.max_row
    last_col_letter = ws.cell(row=1, column=len(filtered_columns)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    _auto_fit_columns(ws)

    wb.save(output_path)
    return output_path


def export_grouped_to_excel(grouped_data: List[dict], output_path: str, group_field_name: str) -> str:
    """
    Экспортирует сгруппированные данные в Excel.
    
    Args:
        grouped_data: Список словарей с полями ["group_name", "count", "cheque_count", "total"]
        output_path: Путь к выходному файлу
        group_field_name: Название поля группировки (category1, category2, etc.)
    
    Returns:
        Путь к сохраненному файлу
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    field_names = {
        "category1": "категория 1 уровня",
        "category2": "категория 2 уровня",
        "category3": "категория 3 уровня",
        "organization": "организация",
        "description": "комментарий"
    }
    
    header_name = field_names.get(group_field_name, group_field_name)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped"
    
    # Заголовки
    headers = [header_name, "количество позиций", "количество чеков", "сумма"]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        _apply_header_style(c)
    
    # Данные
    for r_idx, item in enumerate(grouped_data, start=2):
        ws.cell(row=r_idx, column=1, value=item.get("group_name") or "")
        ws.cell(row=r_idx, column=2, value=item.get("count", 0))
        ws.cell(row=r_idx, column=3, value=item.get("cheque_count", 0))
        total_cell = ws.cell(row=r_idx, column=4, value=float(item.get("total", 0)))
        total_cell.number_format = numbers.FORMAT_NUMBER_00
        _apply_body_style(total_cell, "total")
    
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Autofilter
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:D{last_row}"
    
    _auto_fit_columns(ws)
    
    wb.save(output_path)
    return output_path


def _export_filtered_to_excel(filtered_data: List[dict], output_path: str) -> str:
    """
    Экспортирует отфильтрованные данные в Excel.
    
    Args:
        filtered_data: Список словарей с данными записей
        output_path: Путь к выходному файлу
    
    Returns:
        Путь к сохраненному файлу
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    if not filtered_data:
        raise ValueError("No data to export")
    
    # Получаем колонки из первой записи
    columns = list(filtered_data[0].keys())
    
    # Исключаем из вывода: file_path, created_at, discount, username
    excluded_columns = {"file_path", "created_at", "discount", "username"}
    filtered_columns = [col for col in columns if col not in excluded_columns]
    column_indices = [i for i, col in enumerate(columns) if col not in excluded_columns]
    
    # Перемещаем quantity сразу после product_name
    if "product_name" in filtered_columns and "quantity" in filtered_columns:
        product_idx = filtered_columns.index("product_name")
        quantity_idx = filtered_columns.index("quantity")
        if quantity_idx != product_idx + 1:
            filtered_columns.pop(quantity_idx)
            orig_quantity_idx = column_indices.pop(quantity_idx)
            filtered_columns.insert(product_idx + 1, "quantity")
            column_indices.insert(product_idx + 1, orig_quantity_idx)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchases"
    
    # Header row with RU descriptions
    for col_idx, col_name in enumerate(filtered_columns, start=1):
        header_text = _RU_HEADERS.get(col_name, col_name)
        c = ws.cell(row=1, column=col_idx, value=header_text)
        _apply_header_style(c)
    
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Body rows
    for r_idx, row_dict in enumerate(filtered_data, start=2):
        row_values = [row_dict.get(col) for col in columns]
        for c_idx, orig_idx in enumerate(column_indices, start=1):
            col_name = filtered_columns[c_idx - 1]
            val = row_values[orig_idx]
            coerced = _coerce_cell_value(col_name, val)
            cell = ws.cell(row=r_idx, column=c_idx, value=coerced)
            _apply_body_style(cell, col_name)
    
    # Autofilter
    last_row = ws.max_row
    last_col_letter = ws.cell(row=1, column=len(filtered_columns)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    
    _auto_fit_columns(ws)
    
    wb.save(output_path)
    return output_path

